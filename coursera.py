from random import sample
from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook
from argparse import ArgumentParser
import requests


COURSERA_COURSES_URL = 'https://www.coursera.org/sitemap~www~courses.xml'


def get_web_page(url, payload=None):
    return requests.get(url, payload).text


def get_courses_list(xml_text):
    xml_bytes_text = bytes(xml_text, encoding='utf-8')
    xml = etree.XML(xml_bytes_text)
    return [element.text for element in xml.iter('*') if element.text.rstrip()]


def get_random_elements_from_list(list_elements, count_random_elements=5):
    return sample(list_elements, count_random_elements)


def get_course_info(course_slug):
    soup = BeautifulSoup(course_slug, 'html.parser')
    title = soup.find('h1', class_='title').string
    start_date = soup.find('div', class_='startdate').string
    starting_date = start_date.split(maxsplit=1)[1]
    languages = soup.find('div', class_='language-info').text
    language = languages.split(',')[0]
    duration_in_weeks = len(soup.find_all('div', class_='week'))
    rating_tag = soup.find('div', class_='ratings-text')
    if rating_tag and rating_tag.string:
        rating = rating_tag.string.split()[0]
    else:
        rating = 'No rating'
    return {'title': title,
            'starting_date': starting_date,
            'language': language,
            'duration_in_weeks': duration_in_weeks,
            'rating': rating}


def get_excel_book_with_courses_info(courses_info):
    excel_book = Workbook()
    sheet = excel_book.active
    sheet.title = 'Coursera'
    sheet['A1'] = 'Course title'
    sheet['B1'] = 'Starting date'
    sheet['C1'] = 'Language'
    sheet['D1'] = 'Duration (weeks)'
    sheet['E1'] = 'Rating'
    for row, course in enumerate(courses_info, 2):
        sheet.cell(row=row, column=1, value=course['title'])
        sheet.cell(row=row, column=2, value=course['starting_date'])
        sheet.cell(row=row, column=3, value=course['language'])
        sheet.cell(row=row, column=4, value=course['duration_in_weeks'])
        sheet.cell(row=row, column=5, value=course['rating'])
    return excel_book


def output_courses_info_to_xlsx(excel_book, filepath):
    excel_book.save(filepath)


def get_program_args():
    parser = ArgumentParser(description='Finding coursera courses info')
    parser.add_argument('--count_courses', type=int, default=20,
                        help='Count required courses')
    parser.add_argument('--excel_file', type=str, default='coursera_courses.xlsx',
                        help='Path to output excel file')
    args = parser.parse_args()
    return args.count_courses, args.excel_file


if __name__ == '__main__':
    count_courses, output_excel_file = get_program_args()
    xml_text = get_web_page(COURSERA_COURSES_URL)
    courses_list = get_courses_list(xml_text)
    random_courses_list = get_random_elements_from_list(courses_list,
                                                        count_courses)
    html_pages = [get_web_page(url) for url in random_courses_list]
    coursera_courses_info = [get_course_info(html_page)
                             for html_page in html_pages]
    excel_book = get_excel_book_with_courses_info(coursera_courses_info)
    output_courses_info_to_xlsx(excel_book, output_excel_file)
